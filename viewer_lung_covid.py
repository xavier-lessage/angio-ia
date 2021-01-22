import sys, os
from PyQt5 import QtWidgets
import numpy as np
import pydicom

from PyQt5.uic import loadUiType
from PyQt5 import QtCore

from matplotlib.figure import Figure, Rectangle

from matplotlib.backends.backend_qt4agg import (FigureCanvasQTAgg as FigureCanvas,NavigationToolbar2QT as NavigationToolbar)

from matplotlib import patches

import easygui
from lib_viewer import read_excel, distance_mm

import openpyxl

import json
import keyboard


# Variables globales
listefichiers = []
#SOURCE_EXCEL="C:/Users/xle/Bureau/XL/Privé/Thèse/MammoAI/Tagger/script/COVID.xls"
SOURCE_EXCEL = "C:/Users/xle/Bureau/XL/Privé/Thèse/LungAI/cfg/COVID.xls"

# LOG_FILE="C:/Users/xle/Bureau/XL/Privé/Thèse/MammoAI/Tagger/script/DicomViewer.log"
LOG_FILE= "C:/Users/xle/Bureau/XL/Privé/Thèse/LungAI/cfg/DicomViewer.log"

#PATH_JSON_FILES="C:/Users/xle/Bureau/XL/Privé/Thèse/MammoAI/LungTagger/JSON_FILES"
PATH_JSON_FILES= "C:/Users/xle/Bureau/XL/Privé/Thèse/LungAI/json_files"

PATH_CROPS= "C:/Users/xle/Bureau/XL/Privé/Thèse/LungAI/crops/anomalies"

# SOURCE = "C:/Users/xle/Bureau/XL/Privé/Thèse/MammoAI/Tagger/script/COVID.xls"
SOURCE_EXCEL = "C:/Users/xle/Bureau/XL/Privé/Thèse/LungAI/cfg/COVID.xls"

# DB : C:\Users\xle\Bureau\XL\Privé\Thèse\MammoAI\Tagger\SOURCE_IMG



NBRE_LIGNE_EXCEL = 20

Ui_MainWindow, QMainWindow = loadUiType('tagger_lung.ui')


        
class Main(QMainWindow, Ui_MainWindow):
    def __init__(self, ):
        super(Main, self).__init__()
        self.setupUi(self)
        self.NextImageButton.clicked.connect(self.ShowNextImage)
        self.PreviousImageButton.clicked.connect(self.ShowPreviousImage)
        self.SourceButton.clicked.connect(self.SelectSource)
        self.MesureButton.clicked.connect(self.Mesure)
        self.Anomalie1_Button.clicked.connect(self.SelectAnomalie1)
        self.Anomalie2_Button.clicked.connect(self.SelectAnomalie2)
        self.Anomalie3_Button.clicked.connect(self.SelectAnomalie3)
        self.Mamelon2_Button.clicked.connect(self.SelectMamelon)
        self.ContrasteNormal_Button.clicked.connect(self.ContrasteNormal)
        self.ContrasteEleve_Button.clicked.connect(self.ContrasteEleve)
        self.GotoButton.clicked.connect(self.GoTo)
        self.Segmentation_Button.clicked.connect(self.Segmentation2)
        self.Add_Outline_Button.clicked.connect(self.Add_Outline)
        self.Crop1_Button.clicked.connect(self.Select_Crop1)
        self.Reset_Red_Button.clicked.connect(self.Reset_Select_Red)
        self.Windows_Slider.valueChanged.connect(self.Change_Windows)
        self.Level_Slider.valueChanged.connect(self.Change_Level)
        self.ResetAll_Button.clicked.connect(self.ResetAll_Contour)
        self.Test_Button.clicked.connect(self.Reload_Json_File)

        self.NombreAnomalie = 0
        self.CoordMamelon = []
        self.CoordAnomalie1 = []
        self.CoordAnomalie2 = []
        self.CoordAnomalie3 = []
        self.MammoCourante = ''
        self.CurseurSourceExcel = 1
        self.DataSource = ''
        self.Distance = 0
        self.CoefPxMM = 1
        self.Contraste = 1
        self.constrast_windows = 100

        self.select_temporaire = []
        self.select_y = [0]
        self.select_o = [0]
        self.select_r = [0]
        self.line_color = 'blue'

    def ResetAll_Contour(self,fig):
        self.canvas.figure.gca().clear()
        dataset = pydicom.dcmread(self.NomDuFichierLabel.text())
        self.Reload_Json_File(self)

        level = 50 # self.Level_Slider.value()  # 750  # dataset[0x0028, 0x1051][1]
        window = 750 #self.Windows_Slider.value()
        vmin = level - (window / self.Contraste)
        vmax = level + (window / self.Contraste)
        self.CoefPxMM = 1  # int(dataset[0x0018, 0x7026][0])        self.canvas.figure.gca().clear()
        self.canvas.figure.gca().imshow(dataset.pixel_array, cmap="gray", vmin=vmin, vmax=vmax)
        self.canvas.figure.gca().axis('off')
        self.canvas.figure.subplots_adjust(left=0, right=1, top=1, bottom=0, wspace=0, hspace=0)
        self.canvas.figure.patch.set_facecolor('xkcd:black')

        # Rechargement des contours
        """
        k = 1
        #for k in range(1,len(self.select_r)-4):
        while k < (len(self.select_r)-5):

            if k == 1:
                x0 = self.select_r[k]
                y0 = self.select_r[k + 1]

            x1 = self.select_r[k]
            y1 = self.select_r[k + 1]
            x2 = self.select_r[k + 2]
            y2 = self.select_r[k + 3]
            k = k + 2

            self.line_color = 'red'
            self.canvas.figure.gca().add_artist(patches.mlines.Line2D([x1, x2], [y1, y2], color=self.line_color, linestyle='solid', linewidth=0.1))

            if self.select_r[k+4] == 0:
                self.line_color = 'yellow'
                self.canvas.figure.gca().add_artist(
                patches.mlines.Line2D([x2, x0], [y2, y0], color=self.line_color, linestyle='solid', linewidth=0.1))

        """
        self.canvas.draw()







    def Change_Level(self,fig):
            self.canvas.figure.gca().clear()
            dataset = pydicom.dcmread(self.NomDuFichierLabel.text())

            level = self.Level_Slider.value() #750  # dataset[0x0028, 0x1051][1]
            window = self.Windows_Slider.value()

            vmin = level - (window / self.Contraste)
            vmax = level + (window / self.Contraste)

            self.CoefPxMM = 1  # int(dataset[0x0018, 0x7026][0])        self.canvas.figure.gca().clear()
            self.canvas.figure.gca().imshow(dataset.pixel_array, cmap="gray", vmin=vmin, vmax=vmax)

            self.canvas.figure.gca().axis('off')
            self.canvas.figure.subplots_adjust(left=0, right=1, top=1, bottom=0, wspace=0, hspace=0)
            self.canvas.figure.patch.set_facecolor('xkcd:black')

            self.canvas.draw()

    def Change_Windows(self,fig):
            self.canvas.figure.gca().clear()
            dataset = pydicom.dcmread(self.NomDuFichierLabel.text())

            # if level != 50:  # 2424 #dataset[0x0028, 0x1050][1]
            level = self.Level_Slider.value()
            window = self.Windows_Slider.value() #750  # dataset[0x0028, 0x1051][1]

            vmin = level - (window / self.Contraste)
            vmax = level + (window / self.Contraste)

            self.CoefPxMM = 1  # int(dataset[0x0018, 0x7026][0])        self.canvas.figure.gca().clear()
            self.canvas.figure.gca().imshow(dataset.pixel_array, cmap="gray", vmin=vmin, vmax=vmax)

            self.canvas.figure.gca().axis('off')
            self.canvas.figure.subplots_adjust(left=0, right=1, top=1, bottom=0, wspace=0, hspace=0)
            self.canvas.figure.patch.set_facecolor('xkcd:black')

            self.canvas.draw()

    def Reset_Select_Red(self,fig):
        #self.canvas = FigureCanvas(fig)
        #self.mplvl.addWidget(self.canvas)
        #self.canvas.draw()
        #self.toolbar = NavigationToolbar(self.canvas, self.mplwindow, coordinates=True)
        #self.mplvl.addWidget(self.toolbar)
        #image_limit = fig.get_axes()
        #ym = str(image_limit[0].viewLim.x0)
        self.canvas.figure.gca().clear()
        self.canvas.draw()

    def Reload_Json_File(self,fig):
        a = 12
        CT_WithExtension = os.path.basename(self.NomDuFichierLabel.text())
        CT_WithoutExtension = os.path.splitext(CT_WithExtension)[0]
        JSON_FILE = PATH_JSON_FILES + '/' + CT_WithoutExtension + '.json'
        # lecture du fichier json les coordonnées

        if os.path.exists(JSON_FILE):  # lecture du log file de l'image en cours

            with open(JSON_FILE) as json_data:
                data_dict = json.load(json_data)
                print(data_dict)

            import ast

            self.select_r = ast.literal_eval(data_dict['RedOutlines'])
            self.select_o = ast.literal_eval(data_dict['OrangeOutlines'])
            self.select_y = ast.literal_eval(data_dict['YellowOutlines'])



    def Select_Crop1(self):

        if self.Crop1_button.text() == 'R 1 - OK':

            self.CoordAnomalie1 = ['0', '0', '0', '0']
            self.Crop1_Button.setText('R 1')

        else:

            x1 = self.canvas.figure.axes[0].viewLim._points[0, 0]
            y1 = self.canvas.figure.axes[0].viewLim._points[1, 1]
            y2 = self.canvas.figure.axes[0].viewLim._points[0, 1]
            x2 = self.canvas.figure.axes[0].viewLim._points[1, 0]

            self.canvas.figure.gca().add_artist(patches.Rectangle((x1, y1), x2 - x1, y2 - y1, edgecolor='red', fill=False, linestyle='dashed',
                                      linewidth=0.3, zorder=1))
            self.canvas.figure.gca().text(x1, y1, 'A1', horizontalalignment='left', verticalalignment='top',
                                              color='red', size='3')
            self.CoordAnomalie1 = [x1, x2, y1, y2]

            self.Crop1_Button.setText('R 1 - OK')


    def mouse_move(self,event):
        x, y = event.xdata, event.ydata
        #self.xlxl_label.setText(str(x))

        keyboard_temp = False
        keyboard_y = False
        keyboard_o = False
        keyboard_r = False



        if keyboard.is_pressed('r') or  keyboard.is_pressed('o') or  keyboard.is_pressed('y'):
            keyboard_temp = True

            #self.canvas.figure.gca().scatter(x, y, c='red', s=0.1)
            self.select_temporaire.append(round(x,2))
            self.select_temporaire.append(round(y,2))

            x0 = round(self.select_temporaire[0],2)
            y0 = round(self.select_temporaire[1],2)

            if len(self.select_temporaire) == 2:
                x1 = x0
                x2 = x0
                y1 = y0
                y2 = y0

            else:
                x1 = round(self.select_temporaire[self.k],2)
                y1 = round(self.select_temporaire[self.k + 1],2)
                x2 = round(self.select_temporaire[self.k + 2],2)
                y2 = round(self.select_temporaire[self.k + 3],2)
                self.k = self.k + 2
                if keyboard.is_pressed('r'):
                    self.line_color = 'red'
                    self.canvas.figure.gca().add_artist(patches.mlines.Line2D([x1, x2], [y1, y2], color=self.line_color, linestyle='solid', linewidth=0.1))
                    self.canvas.draw()
                if keyboard.is_pressed('o'):
                    self.line_color = 'orange'
                    self.canvas.figure.gca().add_artist(patches.mlines.Line2D([x1, x2], [y1, y2], color=self.line_color, linestyle='solid', linewidth=0.1))
                    self.canvas.draw()
                if keyboard.is_pressed('y'):
                    self.line_color = 'yellow'
                    self.canvas.figure.gca().add_artist(patches.mlines.Line2D([x1, x2], [y1, y2], color=self.line_color, linestyle='solid', linewidth=0.1))
                    self.canvas.draw()

        if ((len(self.select_temporaire)) >= 2) and (keyboard_temp == False):    # Tracé du dernier segment
            x0 = round(self.select_temporaire[0],2)
            y0 = round(self.select_temporaire[1],2)
            x2 = round(self.select_temporaire[self.k],2)
            y2 = round(self.select_temporaire[self.k + 1],2)
            self.canvas.figure.gca().add_artist(patches.mlines.Line2D([x2, x0], [y2, y0], color=self.line_color, linestyle='solid', linewidth=0.1))
            self.canvas.draw()

            if self.line_color == 'yellow':
                self.select_y = self.select_y + self.select_temporaire
                self.select_y.append(0)

            if self.line_color == 'orange':
                self.select_o = self.select_o + self.select_temporaire
                self.select_o.append(0)

            if self.line_color == 'red':
                self.select_r = self.select_r + self.select_temporaire
                self.select_r.append(0)


            self.select_temporaire = []
            self.k = 0

            CT_WithExtension = os.path.basename(self.NomDuFichierLabel.text())
            CT_WithoutExtension = os.path.splitext(CT_WithExtension)[0]
            JSON_FILE = PATH_JSON_FILES + '/' + CT_WithoutExtension + '.json'

            if self.line_color == 'yellow' or self.line_color == 'orange' or self.line_color == 'red':

                if os.path.exists(JSON_FILE):
                    os.remove(JSON_FILE)

                json_file = open(JSON_FILE, "w")

                json_file.write('{' + '\n')
                
                json_file.write('    "FileName":' + ' ' + '"' + self.NomDuFichierLabel.text() + '",' + '\n')
                json_file.write('    "YellowOutlines":' + ' ' + '"' + str(self.select_y) + '",' + '\n')
                json_file.write('    "OrangeOutlines":' + ' ' + '"' + str(self.select_o) + '",' + '\n')
                json_file.write('    "RedOutlines":' + ' ' + '"' + str(self.select_r) + '"' + '\n')

                json_file.write('}')

                json_file.close()











       # self.canvas.draw()

    def Add_Outline(self):
        self.select_temporaire = []
        self.k = 0
        self.canvas.mpl_connect('motion_notify_event', self.mouse_move)


    def Segmentation2(self):
        self.select_temporaire = []
        #self.xlxl_label.setText('bbb')
      #  self.canvas.figure.gca().axis('on')
       # self.canvas.figure.subplots_adjust(left=0, right=1, top=1, bottom=0, wspace=0, hspace=0)
       # self.canvas.figure.patch.set_facecolor('xkcd:white')
        self.canvas.mpl_connect('motion_notify_event', self.mouse_move)
       # self.canvas.draw()





    def Segmentation(self):
        # self.canvas.figure.cursor(useblit=True, color='red', linewidth=1)
        pts = []
        pts = np.asarray(self.canvas.figure.ginput(7, timeout=-1))

        xm1 = pts[0][0]
        ym1 = pts[0][1]
        xm2 = pts[1][0]
        ym2 = pts[1][1]

        xm3 = pts[2][0]
        ym3 = pts[2][1]
        xm4 = pts[3][0]
        ym4 = pts[3][1]

        xm4 = pts[4][0]
        ym4 = pts[4][1]
        xm5 = pts[5][0]
        ym5 = pts[5][1]

        xm6 = pts[6][0]
        ym6 = pts[6][1]

        self.canvas.figure.gca().add_artist(
            patches.mlines.Line2D([xm1, xm2], [ym1, ym2], color='red', linestyle='dotted', linewidth=0.3))
        self.canvas.figure.gca().add_artist(
            patches.mlines.Line2D([xm2, xm3], [ym2, ym3], color='red', linestyle='dotted', linewidth=0.3))
        self.canvas.figure.gca().add_artist(
            patches.mlines.Line2D([xm3, xm4], [ym3, ym4], color='red', linestyle='dotted', linewidth=0.3))
        self.canvas.figure.gca().add_artist(
            patches.mlines.Line2D([xm4, xm5], [ym4, ym5], color='red', linestyle='dotted', linewidth=0.3))
        self.canvas.figure.gca().add_artist(
            patches.mlines.Line2D([xm5, xm6], [ym5, ym6], color='red', linestyle='dotted', linewidth=0.3))
        # self.canvas.figure.gca().add_artist(patches.mlines.Line2D([xm6, xm7], [ym6, ym7], color='red', linestyle='dotted', linewidth=0.3))
        self.canvas.draw()

    def GoTo(self):
        self.CurseurSourceExcel = int(self.LigneExcel_spinBox.value())
        self.ShowInformations()

    def SelectSource(self):
        SOURCE = easygui.diropenbox()

        self.DataSource = SOURCE
        self.DataSource = self.DataSource.replace('\\', '/')

        for path, subdirs, files in os.walk(SOURCE):
            for name in files:
                f = os.path.join(path, name)
                listefichiers.append(f)

        for i in range(len(listefichiers)):
            listefichiers[i] = listefichiers[i].replace('\\', '/')

        self.NomDuFichierLabel.setText(listefichiers[0])
        self.MammoCourante = listefichiers[0]
        #self.ShowInformations()
        #self.LigneExcellabel.setText(str(self.CurseurSourceExcel))

     #   self.AnomaliesButton.setEnabled(True)
     #   self.MamelonButton.setEnabled(True)
   #     self.ResetButton.setEnabled(True)
        self.MesureButton.setEnabled(True)
        self.PreviousImageButton.setEnabled(True)
        self.NextImageButton.setEnabled(True)

        # Affiche la dernière image traitée
        # Placer le code de Salvatore ici et extraire le numéro de la ligne Excel

        if os.path.exists(LOG_FILE):    # lecture du log file avec la dernière image traitée
            json_file = open(LOG_FILE, "r")
            LigneJsonFile = json_file.readlines()
            LastFileProcesed = LigneJsonFile[0][18-len(LigneJsonFile)]    # Récupération du numéro de la ligne
            json_file.close()
            self.CurseurSourceExcel = int(LastFileProcesed)     # Mise à jour du curseur

        else:    # Initialisation du log file si pas présent
            json_file = open(LOG_FILE, "w")
            json_file.write('LastFileProcesed=2')
            json_file.close()

        #self.LigneExcellabel.setText(str(self.CurseurSourceExcel))
        self.ShowInformations()



    def loadImage(self, fig):
        self.canvas = FigureCanvas(fig)
        self.mplvl.addWidget(self.canvas)
        self.canvas.draw()
        self.toolbar = NavigationToolbar(self.canvas,self.mplwindow, coordinates=True)
        self.mplvl.addWidget(self.toolbar)
        image_limit=fig.get_axes()
        ym = str(image_limit[0].viewLim.x0)

        self.canvas.figure.gca().clear()

     #   self.AnomaliesButton.setEnabled(False)
    #    self.MamelonButton.setEnabled(False)
    #    self.ResetButton.setEnabled(False)
        self.MesureButton.setEnabled(False)
        self.PreviousImageButton.setEnabled(False)
        self.NextImageButton.setEnabled(False)

        self.anomalie1 = False
        self.anomalie2 = False
        self.anomalie3 = False

        self.ContrasteNormal_Button.setChecked(True)

        # Pour le contour
        self.select_temporaire = []
        self.k = 0
        self.canvas.mpl_connect('motion_notify_event', self.mouse_move)



    def ContrasteNormal(self):
        self.Contraste = 1
   #     self.ContrasteNormal_Button.setChecked(True)
    #    self.ContrasteEleve_Button.setChecked(False)
        self.ShowInformations()


    def ContrasteEleve(self):
        self.Contraste = 2
     #   self.ContrasteNormal_Button.setChecked(False)
     #   self.ContrasteEleve_Button.setChecked(True)
        self.ShowInformations()


    def ShowInformations(self):
        # Initialisation des flags
        self.CoordMamelon = []
        self.CoordAnomalie1 = []
        self.CoordAnomalie2 = []
        self.CoordAnomalie3 = []

        self.AnomaliesVisiblescheckBox.setChecked(False)
        self.radioDroitButton.setChecked(False)
        self.radioAButton.setChecked(False)
        self.MassecheckBox.setChecked(False)
        self.MicroCalcificationscheckBox.setChecked(False)
        self.DesorganisationcheckBox.setChecked(False)
        self.AsymetriecheckBox.setChecked(False)
        self.retromamcheckBox.setChecked(False)
        self.SuperieurcheckBox.setChecked(False)
        self.InferieurcheckBox.setChecked(False)
        self.ExternecheckBox.setChecked(False)
        self.InternecheckBox.setChecked(False)

        # Lecture des informations à partir du fichier source Excel

        Infos=read_excel(SOURCE_EXCEL)
        LigneCouranteFichierSource=self.CurseurSourceExcel
        self.LigneExcel_spinBox.setValue(int(LigneCouranteFichierSource))
        self.LigneExcel_spinBox.setMaximum(int(NBRE_LIGNE_EXCEL)-1)

        self.MesureButton.setEnabled(False)

        if (Infos[LigneCouranteFichierSource][33] == '0.0') & (Infos[LigneCouranteFichierSource][21] == '0.0') & (Infos[LigneCouranteFichierSource][25] == '0.0') & (Infos[LigneCouranteFichierSource][29] == '0.0'):
   #         self.MamelonButton.setEnabled(True)
    #        self.AnomaliesButton.setEnabled(True)
            self.MesureButton.setEnabled(True)

        # Si anomalie 1 présente, mettre le statut A1 - OK
        if Infos[LigneCouranteFichierSource][21] == '0.0':
            self.Anomalie1_Button.setText('Anomalie 1')
            self.CoordAnomalie1 = ['0', '0', '0', '0']
        else:
            self.Anomalie1_Button.setText('A1 - OK')

        # Si anomalie 2 présente, mettre le statut A2 - OK
        if Infos[LigneCouranteFichierSource][25] == '0.0':
            self.Anomalie2_Button.setText('Anomalie 2')
            self.CoordAnomalie2 = ['0', '0', '0', '0']
        else:
            self.Anomalie2_Button.setText('A2 - OK')

        # Si anomalie 3 présente, mettre le statut A3 - OK
        if Infos[LigneCouranteFichierSource][29] == '0.0':
            self.Anomalie3_Button.setText('Anomalie 3')
            self.CoordAnomalie3 = ['0', '0', '0', '0']
        else:
            self.Anomalie3_Button.setText('A3 - OK')

        # Mamelon non présent --> Reset
        if Infos[LigneCouranteFichierSource][33] == '0.0':
            self.Mamelon2_Button.setText('Mamelon')
            self.CoordMamelon = ['0', '0', '0', '0']
        else:
            #self.CoordMamelon = []
            self.Mamelon2_Button.setText('M - OK')

        # Nombre d'anomalies
        self.NombreAnomalie = 3
        if Infos[LigneCouranteFichierSource][21] == '0.0':
            self.NombreAnomalie = self.NombreAnomalie - 1
        if Infos[LigneCouranteFichierSource][25] == '0.0':
            self.NombreAnomalie = self.NombreAnomalie - 1
        if Infos[LigneCouranteFichierSource][29] == '0.0':
            self.NombreAnomalie = self.NombreAnomalie - 1



        # Infos[LigneCouranteFichierSource][4] --> Anomalie(s) visible(s)
        if Infos[LigneCouranteFichierSource][4] == '1.0':
            self.AnomaliesVisiblescheckBox.setChecked(True)
        else:
            self.AnomaliesVisiblescheckBox.setChecked(False)

        # Infos[LigneCouranteFichierSource][6] --> Masse
        if Infos[LigneCouranteFichierSource][6] == '1.0':
            self.MassecheckBox.setChecked(True)
        else:
            self.MassecheckBox.setChecked(False)

        # Infos[LigneCouranteFichierSource][7] --> MicroCalcifications
        if Infos[LigneCouranteFichierSource][7] == '1.0':
            self.MicroCalcificationscheckBox.setChecked(True)
        else:
            self.MicroCalcificationscheckBox.setChecked(False)

        # Infos[LigneCouranteFichierSource][8] --> Désorganisation
        if Infos[LigneCouranteFichierSource][8] == '1.0':
            self.DesorganisationcheckBox.setChecked(True)
        else:
            self.DesorganisationcheckBox.setChecked(False)

        # Infos[LigneCouranteFichierSource][9] --> Asymétrie
        if Infos[LigneCouranteFichierSource][9] == '1.0':
            self.AsymetriecheckBox.setChecked(True)
        else:
            self.AsymetriecheckBox.setChecked(False)

        # Infos[LigneCouranteFichierSource][10] --> Dimension
        if Infos[LigneCouranteFichierSource][10] != "":
            self.Dimensionlabel.setText('Dimension: ' + Infos[LigneCouranteFichierSource][10]+ ' mm')

        # Infos[LigneCouranteFichierSource][13] --> Supérieur
        if Infos[LigneCouranteFichierSource][13] == '1.0':
            self.SuperieurcheckBox.setChecked(True)
        else:
            self.SuperieurcheckBox.setChecked(False)

        # Infos[LigneCouranteFichierSource][14] --> Inferieur
        if Infos[LigneCouranteFichierSource][14] == '1.0':
            self.InferieurcheckBox.setChecked(True)
        else:
            self.InferieurcheckBox.setChecked(False)

        # Infos[LigneCouranteFichierSource][15] --> Externe
        if Infos[LigneCouranteFichierSource][15] == '1.0':
            self.ExternecheckBox.setChecked(True)
        else:
            self.ExternecheckBox.setChecked(False)

        # Infos[LigneCouranteFichierSource][16] --> Interne
        if Infos[LigneCouranteFichierSource][16] == '1.0':
            self.InternecheckBox.setChecked(True)
        else:
            self.InternecheckBox.setChecked(False)

        # Infos[LigneCouranteFichierSource][17] --> retromamcheckBox
        if Infos[LigneCouranteFichierSource][17] == '1.0':
            self.retromamcheckBox.setChecked(True)
        else:
            self.retromamcheckBox.setChecked(False)

        # Infos[LigneCouranteFichierSource][11] --> Côté droit
        if Infos[LigneCouranteFichierSource][11] == 'R':
            self.radioDroitButton.setChecked(True)

        # Infos[LigneCouranteFichierSource][11] --> Côté droit
        if Infos[LigneCouranteFichierSource][11] == 'L':
            self.radioGaucheButton.setChecked(True)

        # Infos[LigneCouranteFichierSource][20] --> Densité
        if Infos[LigneCouranteFichierSource][20] == 'A':
            self.radioAButton.setChecked(True)
        if Infos[LigneCouranteFichierSource][20] == 'B':
            self.radioBButton.setChecked(True)
        if Infos[LigneCouranteFichierSource][20] == 'C':
            self.radioCButton.setChecked(True)
        if Infos[LigneCouranteFichierSource][20] == 'D':
            self.radioDButton.setChecked(True)

        # Affichage de l'image et anomalie(s)s éventuelles
        filename = self.DataSource + Infos[LigneCouranteFichierSource][19]

        self.NomDuFichierLabel.setText(filename)
        self.LigneExcellabel.setText(str(self.CurseurSourceExcel))

        dataset = pydicom.dcmread(filename)
        level = 50 #2424 #dataset[0x0028, 0x1050][1]
        window = 750 #dataset[0x0028, 0x1051][1]

        vmin = level - (window / self.Contraste)
        vmax = level + (window / self.Contraste)

        self.CoefPxMM = 1 #int(dataset[0x0018, 0x7026][0])        self.canvas.figure.gca().clear()
        self.canvas.figure.gca().imshow(dataset.pixel_array, cmap="gray", vmin=vmin, vmax=vmax)

        self.canvas.figure.gca().axis('off')
        self.canvas.figure.subplots_adjust(left=0, right=1, top=1, bottom=0, wspace=0, hspace=0)
        self.canvas.figure.patch.set_facecolor('xkcd:black')

        # Mamelon
        if float(Infos[LigneCouranteFichierSource][33]) > 0:
            x1 = float(Infos[LigneCouranteFichierSource][33])
            x2 = float(Infos[LigneCouranteFichierSource][34])
            y1 = float(Infos[LigneCouranteFichierSource][35])
            y2 = float(Infos[LigneCouranteFichierSource][36])
            self.canvas.figure.gca().add_artist(patches.Rectangle((x1, y1), x2 - x1, y2 - y1, edgecolor='blue', fill=False, linestyle='dashdot', linewidth=0.3,zorder=1))
            self.canvas.figure.gca().text(x1, y1, 'M', horizontalalignment='left', verticalalignment='top',color='blue', size='3')
            self.CoordMamelon = [x1,y1,x2,y2]

        # Anomalie 1
        if float(Infos[LigneCouranteFichierSource][21]) > 0:
            x1 = float(Infos[LigneCouranteFichierSource][21])
            x2 = float(Infos[LigneCouranteFichierSource][22])
            y1 = float(Infos[LigneCouranteFichierSource][23])
            y2 = float(Infos[LigneCouranteFichierSource][24])
            self.canvas.figure.gca().add_artist(patches.Rectangle((x1, y1), x2 - x1, y2 - y1, edgecolor='red', fill=False, linestyle='dashed', linewidth=0.3,zorder=1))
            self.canvas.figure.gca().text(x1, y1, 'A1', horizontalalignment='left', verticalalignment='top',color='red', size='3')
            self.CoordAnomalie1 = [x1, y1, x2, y2]

        # Anomalie 2
        if float(Infos[LigneCouranteFichierSource][25]) > 0:
            x1 = float(Infos[LigneCouranteFichierSource][25])
            x2 = float(Infos[LigneCouranteFichierSource][26])
            y1 = float(Infos[LigneCouranteFichierSource][27])
            y2 = float(Infos[LigneCouranteFichierSource][28])
            self.canvas.figure.gca().add_artist(patches.Rectangle((x1, y1), x2 - x1, y2 - y1, edgecolor='orange', fill=False, linestyle='dashed', linewidth=0.3,zorder=1))
            self.canvas.figure.gca().text(x1, y1, 'A2', horizontalalignment='left', verticalalignment='top',color='orange', size='3')
            self.CoordAnomalie2 = [x1, y1, x2, y2]

        # Anomalie 3
        if float(Infos[LigneCouranteFichierSource][29]) > 0:
            x1 = float(Infos[LigneCouranteFichierSource][29])
            x2 = float(Infos[LigneCouranteFichierSource][30])
            y1 = float(Infos[LigneCouranteFichierSource][31])
            y2 = float(Infos[LigneCouranteFichierSource][32])
            self.canvas.figure.gca().add_artist(patches.Rectangle((x1, y1), x2 - x1, y2 - y1, edgecolor='yellow', fill=False, linestyle='dashed', linewidth=0.3,zorder=1))
            self.canvas.figure.gca().text(x1, y1, 'A3', horizontalalignment='left', verticalalignment='top',color='yellow', size='3')
            self.CoordAnomalie3 = [x1, y1, x2, y2]

        self.canvas.draw()

        self.MesureMGlabel.setText('Mesure MG: ' + str(Infos[LigneCouranteFichierSource][37]) + ' px')


    def ShowNextImage(self,fig):

        #self.Reload_Json_File()
        # disable  self.UpdateValuesExcel()

        # Effacement du contout precedent
        self.ResetAll_Contour(self)
        self.select_y = [0]
        self.select_o = [0]
        self.select_r = [0]

        # Update du fichier log
        # Suppression de l'ancien fichier car déjà présent
        if os.path.exists('LOG_FILE'):
            os.remove(LOG_FILE)

        json_file = open(LOG_FILE, "w")
        if (self.CurseurSourceExcel + 1) == NBRE_LIGNE_EXCEL:
            json_file.write('LastFileProcesed=' + str(self.CurseurSourceExcel))
        else:
             json_file.write('LastFileProcesed=' + str(self.CurseurSourceExcel+1))

        json_file.close()


        #self.debuginfo.setText(str(self.NombreAnomalie))
        if self.CurseurSourceExcel < NBRE_LIGNE_EXCEL - 1:
            self.CurseurSourceExcel = self.CurseurSourceExcel + 1
            #self.CoordMamelon = []
            #self.CoordAnomalie1 = []
            #self.CoordAnomalie2 = []
            #self.CoordAnomalie3 = []
            self.ShowInformations()
            #self.UpdateExcel()



    def ShowPreviousImage(self,fig):
        # disable  self.UpdateValuesExcel()
        #self.debuginfo.setText(str(self.NombreAnomalie))
        # Effacement du contout precedent
        self.ResetAll_Contour(self)
        self.select_y = [0]
        self.select_o = [0]
        self.select_r = [0]

        if self.CurseurSourceExcel >= 2:
            self.CurseurSourceExcel = self.CurseurSourceExcel - 1
            #self.CoordMamelon = []
            #self.CoordAnomalie1 = []
            #self.CoordAnomalie2 = []
            #self.CoordAnomalie3 = []
            self.ShowInformations()
            #self.UpdateExcel()

    def UpdateExcel_Anomalie1(self):
        xfile = openpyxl.load_workbook(SOURCE_EXCEL, read_only = False)
        sheet = xfile['seno']

        # rect1X1
        sheet['V' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie1[0])
        # rect1Y1
        sheet['W' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie1[1])
        # rect1X2
        sheet['X' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie1[2])
        # rect1Y2
        sheet['Y' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie1[3])

        sheet['AL' + str(self.CurseurSourceExcel + 1)] = self.Distance
        self.MesureMGlabel.setText('Mesure MG: ' + str(int(self.Distance)) + ' px')
        xfile.save(SOURCE_EXCEL)


    def UpdateExcel_Anomalie2(self):
        xfile = openpyxl.load_workbook(SOURCE_EXCEL, read_only = False)
        sheet = xfile['seno']

        # rect2X1
        sheet['Z' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie2[0])
        # rect2Y1
        sheet['AA' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie2[1])
        # rect2X2
        sheet['AB' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie2[2])
        # rect2Y2
        sheet['AC' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie2[3])

        sheet['AL' + str(self.CurseurSourceExcel + 1)] = self.Distance
        self.MesureMGlabel.setText('Mesure MG: ' + str(int(self.Distance)) + ' px')
        xfile.save(SOURCE_EXCEL)

    def UpdateExcel_Anomalie3(self):
        xfile = openpyxl.load_workbook(SOURCE_EXCEL, read_only = False)
        sheet = xfile['seno']

        # rect3X1
        sheet['AD' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie3[0])
        # rect3Y1
        sheet['AE' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie3[1])
        # rect3X2
        sheet['AF' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie3[2])
        # rect3Y2
        sheet['AG' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie3[3])

        sheet['AL' + str(self.CurseurSourceExcel + 1)] = self.Distance
        self.MesureMGlabel.setText('Mesure MG: ' + str(int(self.Distance)) + ' px')
        xfile.save(SOURCE_EXCEL)

    def UpdateExcel_Mamelon(self):
        xfile = openpyxl.load_workbook(SOURCE_EXCEL, read_only = False)
        sheet = xfile['seno']

        # mamX1
        sheet['AH' + str(self.CurseurSourceExcel + 1)] = int(self.CoordMamelon[0])
        # mamY1
        sheet['AI' + str(self.CurseurSourceExcel + 1)] = int(self.CoordMamelon[1])
        # mamX2
        sheet['AJ' + str(self.CurseurSourceExcel + 1)] = int(self.CoordMamelon[2])
        # mamY2
        sheet['AK' + str(self.CurseurSourceExcel + 1)] = int(self.CoordMamelon[3])

        sheet['AL' + str(self.CurseurSourceExcel + 1)] = self.Distance
        self.MesureMGlabel.setText('Mesure MG: ' + str(int(self.Distance)) + ' px')
        xfile.save(SOURCE_EXCEL)





    def UpdateExcel2(self):
        xfile = openpyxl.load_workbook(SOURCE_EXCEL, read_only = False)
        sheet = xfile['seno']

        if self.Anomalie1_Button.text() == 'A1 - OK':
            # rect1X1
            sheet['V' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie1[0])
            # rect1Y1
            sheet['W' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie1[1])
            # rect1X2
            sheet['X' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie1[2])
            # rect1Y2
            sheet['Y' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie1[3])

        if self.Anomalie2_Button.text() == 'A2 - OK':
            # rect2X1
            sheet['Z' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie2[0])
            # rect2Y1
            sheet['AA' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie2[1])
            # rect2X2
            sheet['AB' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie2[2])
            # rect2Y2
            sheet['AC' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie2[3])

        if self.Anomalie3_Button.text() == 'A3 - OK':
            # rect3X1
            sheet['AD' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie3[0])
            # rect3Y1
            sheet['AE' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie3[1])
            # rect3X2
            sheet['AF' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie3[2])
            # rect3Y2
            sheet['AG' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie3[3])

        sheet['AL' + str(self.CurseurSourceExcel + 1)] = self.Distance
        self.MesureMGlabel.setText('Mesure MG: ' + str(int(self.Distance)) + ' px')
        xfile.save(SOURCE_EXCEL)



    def SelectAnomalie1(self):

        if self.Anomalie1_Button.text() == 'A1 - OK':
            # Reset de la selection
            self.canvas.figure.gca().clear()
            self.CoordAnomalie1 = ['0', '0', '0', '0']
            # disable  self.UpdateExcel_Anomalie1()
            self.ShowInformations()
            #self.canvas.draw()
            self.Anomalie1_Button.setText('Anomalie 1')

        else:

            # Création du fichier avec l'anomalie 1
            CropNameWithExtension = os.path.basename(self.NomDuFichierLabel.text())
            CropNameWithoutExtension = os.path.splitext(CropNameWithExtension)[0]

            # Suppression de l'ancien fichier si déjà présent
            if os.path.exists(PATH_CROPS + '/' + CropNameWithoutExtension + '_Anomalie1.png'):
                os.remove(PATH_CROPS + '/' + CropNameWithoutExtension + '_Anomalie1.png')
            # Sauvegarde
            self.canvas.figure.savefig(PATH_CROPS + '/' + CropNameWithoutExtension + '_Anomalie1.png', dpi=300)


            # Encadrement de l'anomalie et ajout d'une etiquette
            x1 = self.canvas.figure.axes[0].viewLim._points[0, 0]
            y1 = self.canvas.figure.axes[0].viewLim._points[1, 1]
            y2 = self.canvas.figure.axes[0].viewLim._points[0, 1]
            x2 = self.canvas.figure.axes[0].viewLim._points[1, 0]

            self.canvas.figure.gca().add_artist(patches.Rectangle((x1, y1), x2 - x1, y2 - y1, edgecolor='red', fill=False, linestyle='dashed',linewidth=0.3,zorder=1))
            self.canvas.figure.gca().text(x1, y1, 'A1', horizontalalignment='left', verticalalignment='top', color='red',size='3')
            self.CoordAnomalie1 = [x1, x2, y1, y2]

            self.Anomalie1_Button.setText('A1 - OK')

            # Sauvegarde des coordonnées dans le fichier Excel
            # disable self.UpdateExcel_Anomalie1()

    def SelectAnomalie2(self):

        if self.Anomalie2_Button.text() == 'A2 - OK':
            # Reset de la selection
            self.canvas.figure.gca().clear()
            self.CoordAnomalie2 = ['0', '0', '0', '0']
            # disable  self.UpdateExcel_Anomalie2()
            self.ShowInformations()
            #self.canvas.draw()
            self.Anomalie2_Button.setText('Anomalie 2')

        else:

            # Création du fichier avec l'anomalie 2
            CropNameWithExtension = os.path.basename(self.NomDuFichierLabel.text())
            CropNameWithoutExtension = os.path.splitext(CropNameWithExtension)[0]

            # Suppression de l'ancien fichier si déjà présent
            if os.path.exists(PATH_CROPS + '/' + CropNameWithoutExtension + '_Anomalie2.png'):
                os.remove(PATH_CROPS + '/' + CropNameWithoutExtension + '_Anomalie2.png')
            # Sauvegarde
            self.canvas.figure.savefig(PATH_CROPS + '/' + CropNameWithoutExtension + '_Anomalie2.png', dpi=300)


            # Encadrement de l'anomalie et ajout d'une etiquette
            x1 = self.canvas.figure.axes[0].viewLim._points[0, 0]
            y1 = self.canvas.figure.axes[0].viewLim._points[1, 1]
            y2 = self.canvas.figure.axes[0].viewLim._points[0, 1]
            x2 = self.canvas.figure.axes[0].viewLim._points[1, 0]

            self.canvas.figure.gca().add_artist(patches.Rectangle((x1, y1), x2 - x1, y2 - y1, edgecolor='orange', fill=False, linestyle='dashed',linewidth=0.3,zorder=1))
            self.canvas.figure.gca().text(x1, y1, 'A2', horizontalalignment='left', verticalalignment='top', color='orange',size='3')
            self.CoordAnomalie2 = [x1, x2, y1, y2]

            self.Anomalie2_Button.setText('A2 - OK')

            # Sauvegarde des coordonnées dans le fichier Excel
            # disable self.UpdateExcel_Anomalie2()


    def SelectAnomalie3(self):

        if self.Anomalie3_Button.text() == 'A3 - OK':
            # Reset de la selection
            self.canvas.figure.gca().clear()
            self.CoordAnomalie3 = ['0', '0', '0', '0']
            # disable self.UpdateExcel_Anomalie3()
            self.ShowInformations()
            #self.canvas.draw()
            self.Anomalie3_Button.setText('Anomalie 3')

        else:

            # Création du fichier avec l'anomalie 3
            CropNameWithExtension = os.path.basename(self.NomDuFichierLabel.text())
            CropNameWithoutExtension = os.path.splitext(CropNameWithExtension)[0]

            # Suppression de l'ancien fichier si déjà présent
            if os.path.exists(PATH_CROPS + '/' + CropNameWithoutExtension + '_Anomalie3.png'):
                os.remove(PATH_CROPS + '/' + CropNameWithoutExtension + '_Anomalie3.png')
            # Sauvegarde
            self.canvas.figure.savefig(PATH_CROPS + '/' + CropNameWithoutExtension + '_Anomalie3.png', dpi=300)


            # Encadrement de l'anomalie et ajout d'une etiquette
            x1 = self.canvas.figure.axes[0].viewLim._points[0, 0]
            y1 = self.canvas.figure.axes[0].viewLim._points[1, 1]
            y2 = self.canvas.figure.axes[0].viewLim._points[0, 1]
            x2 = self.canvas.figure.axes[0].viewLim._points[1, 0]

            self.canvas.figure.gca().add_artist(patches.Rectangle((x1, y1), x2 - x1, y2 - y1, edgecolor='yellow', fill=False, linestyle='dashed',linewidth=0.3,zorder=1))
            self.canvas.figure.gca().text(x1, y1, 'A3', horizontalalignment='left', verticalalignment='top', color='yellow',size='3')
            self.CoordAnomalie3 = [x1, x2, y1, y2]

            self.Anomalie3_Button.setText('A3 - OK')

            # Sauvegarde des coordonnées dans le fichier Excel
            # disable self.UpdateExcel_Anomalie3()



    def SelectMamelon(self):

        if self.Mamelon2_Button.text() == 'M - OK':
            # Reset de la selection
            self.canvas.figure.gca().clear()
            self.CoordMamelon = ['0', '0', '0', '0']
            self.UpdateExcel_Mamelon()
            self.ShowInformations()
            #self.canvas.draw()
            self.Mamelon2_Button.setText('Mamelon')

        else:

            # Création du fichier avec le mamelon
            CropNameWithExtension = os.path.basename(self.NomDuFichierLabel.text())
            CropNameWithoutExtension = os.path.splitext(CropNameWithExtension)[0]

            # Suppression de l'ancien fichier si déjà présent
            if os.path.exists('CROPS/Mamelons/' + CropNameWithoutExtension + '_Mamelon.png'):
                os.remove('CROPS/Mamelons/' + CropNameWithoutExtension + '_Mamelon.png')
            # Sauvegarde
            self.canvas.figure.savefig('CROPS/Mamelons/' + CropNameWithoutExtension + '_Mamelon.png', dpi=300)

            # Encadrement du mamelon et ajout d'une etiquette
            x1 = self.canvas.figure.axes[0].viewLim._points[0, 0]
            y1 = self.canvas.figure.axes[0].viewLim._points[1, 1]
            y2 = self.canvas.figure.axes[0].viewLim._points[0, 1]
            x2 = self.canvas.figure.axes[0].viewLim._points[1, 0]

            self.canvas.figure.gca().add_artist(patches.Rectangle((x1, y1), x2 - x1, y2 - y1, edgecolor='blue', fill=False, linestyle='dashed',linewidth=0.3,zorder=1))
            self.canvas.figure.gca().text(x1, y1, 'M', horizontalalignment='left', verticalalignment='top', color='blue',size='3')
            self.CoordMamelon = [x1, x2, y1, y2]

            self.Mamelon2_Button.setText('M - OK')

            # Sauvegarde des coordonnées dans le fichier Excel
            self.UpdateExcel_Mamelon()




    def AnomaliesCreation(self):
        self.NombreAnomalie = self.NombreAnomalie + 1
        x1=self.canvas.figure.axes[0].viewLim._points[0,0]
        y1 = self.canvas.figure.axes[0].viewLim._points[1, 1]
        y2=self.canvas.figure.axes[0].viewLim._points[0,1]
        x2 = self.canvas.figure.axes[0].viewLim._points[1,0]

        if self.NombreAnomalie <= 3:

            if self.CoordAnomalie1 == []:
                # Sauvegarde de l'anomalie 1
                CropNameWithExtension = os.path.basename(self.NomDuFichierLabel.text())
                CropNameWithoutExtension = os.path.splitext(CropNameWithExtension)[0]

                # Suppression de l'ancien fichier si déjà présent
                if os.path.exists('CROPS/Anomalies/' + CropNameWithoutExtension + '_Anomalie1.png'):
                    os.remove('CROPS/Anomalies/' + CropNameWithoutExtension + '_Anomalie1.png')
                # Sauvegarde
                self.canvas.figure.savefig('CROPS/Anomalies/' + CropNameWithoutExtension + '_Anomalie1.png', dpi=300)

                # Encadrement de l'anomalie et ajout d'une etiquette
                self.canvas.figure.gca().add_artist(patches.Rectangle((x1, y1), x2-x1-20, y2-y1-20,edgecolor='red',fill=False, linestyle='dashed',linewidth=0.3, zorder=1))
                self.canvas.figure.gca().text(x1, y1, 'A1', horizontalalignment='left', verticalalignment='top',color='red', size='3')
                self.CoordAnomalie1 = [x1, x2, y1, y2]

                # Sauvegarde des coordonnées dans le fichier Excel
                self.UpdateExcel()

            elif self.CoordAnomalie2 == []:
                # Sauvegarde de l'anomalie 2
                CropNameWithExtension = os.path.basename(self.NomDuFichierLabel.text())
                CropNameWithoutExtension = os.path.splitext(CropNameWithExtension)[0]

                # Suppression de l'ancien fichier si déjà présent
                if os.path.exists('CROPS/Anomalies/' + CropNameWithoutExtension + '_Anomalie2.png'):
                    os.remove('CROPS/Anomalies/' + CropNameWithoutExtension + '_Anomalie2.png')
                # Sauvegarde
                self.canvas.figure.savefig('CROPS/Anomalies/' + CropNameWithoutExtension + '_Anomalie2.png', dpi=300)

                # Encadrement de l'anomalie et ajout d'une etiquette
                self.canvas.figure.gca().add_artist(
                    patches.Rectangle((x1, y1), x2 - x1, y2 - y1, edgecolor='orange', fill=False,linestyle='dashed', linewidth=0.3, zorder=1))
                self.canvas.figure.gca().text(x1, y1, 'A2', horizontalalignment='left', verticalalignment='top',color='orange', size='3')
                self.CoordAnomalie2 = [x1, x2, y1, y2]

                # Sauvegarde des coordonnées dans le fichier Excel
                self.UpdateExcel()

                #self.canvas.figure.gca().add_artist(patches.Rectangle((x1, y1), x2 - x1 - 20, y2 - y1 - 20, edgecolor='red', fill=False,linestyle='dashed', linewidth=0.3, zorder=1))
                #self.canvas.figure.gca().text(x1, y1, 'A2', horizontalalignment='left', verticalalignment='top',color='red',size='3')
                #self.CoordAnomalie2 = [x1, x2, y1, y2]
                #self.UpdateExcel()

            elif self.CoordAnomalie3 == []:
                # Sauvegarde de l'anomalie 3
                CropNameWithExtension = os.path.basename(self.NomDuFichierLabel.text())
                CropNameWithoutExtension = os.path.splitext(CropNameWithExtension)[0]

                # Suppression de l'ancien fichier si déjà présent
                if os.path.exists('CROPS/Anomalies/' + CropNameWithoutExtension + '_Anomalie3.png'):
                    os.remove('CROPS/Anomalies/' + CropNameWithoutExtension + '_Anomalie3.png')
                # Sauvegarde
                self.canvas.figure.savefig('CROPS/Anomalies/' + CropNameWithoutExtension + '_Anomalie3.png', dpi=300)

                # Encadrement de l'anomalie et ajout d'une etiquette
                self.canvas.figure.gca().add_artist(patches.Rectangle((x1, y1), x2 - x1, y2 - y1, edgecolor='yellow', fill=False, linestyle='dashed',linewidth=0.3, zorder=1))
                self.canvas.figure.gca().text(x1, y1, 'A3', horizontalalignment='left', verticalalignment='top',color='yellow', size='3')
                self.CoordAnomalie3 = [x1, x2, y1, y2]

                # Sauvegarde des coordonnées dans le fichier Excel
                self.UpdateExcel()

                # self.canvas.figure.gca().add_artist(patches.Rectangle((x1, y1), x2 - x1 - 20, y2 - y1 - 20, edgecolor='red', fill=False,linestyle='dashed', linewidth=0.3, zorder=1))
                # self.canvas.figure.gca().text(x1, y1, 'A3', horizontalalignment='left', verticalalignment='top',color='red',size='3')
                # self.CoordAnomalie3 = [x1, x2, y1, y2]
                # self.UpdateExcel()


    def MamelonCreation(self):
        x1 = self.canvas.figure.axes[0].viewLim._points[0, 0]
        y1 = self.canvas.figure.axes[0].viewLim._points[1, 1]
        y2 = self.canvas.figure.axes[0].viewLim._points[0, 1]
        x2 = self.canvas.figure.axes[0].viewLim._points[1, 0]

        if self.CoordMamelon == []:
            # Sauvegarde du mamelon
            CropNameWithExtension = os.path.basename(self.NomDuFichierLabel.text())
            CropNameWithoutExtension = os.path.splitext(CropNameWithExtension)[0]
            # Suppression de l'ancien fichier si déjà présent
            if os.path.exists('CROPS/Mamelons/' + CropNameWithoutExtension + '_Mamelon.png'):
                os.remove('CROPS/Mamelons/' + CropNameWithoutExtension + '_Mamelon.png')
            # Sauvegarde
            self.canvas.figure.savefig('CROPS/Mamelons/' + CropNameWithoutExtension + '_Mamelon.png', dpi=300)
            # Encadrement du mamelon et ajout d'une etiquette
            self.canvas.figure.gca().add_artist(patches.Rectangle((x1, y1), x2 - x1, y2 - y1, edgecolor='blue', fill=False, linestyle='dashdot',linewidth=0.3, zorder=1))
            self.canvas.figure.gca().text(x1, y1, 'M', horizontalalignment='left', verticalalignment='top',color='blue', size='3')
            self.CoordMamelon = [x1, x2, y1, y2]
            # Sauvegarde des coordonnées dans le fichier Excel
            self.UpdateExcel()


    def Mesure(self):
        #self.canvas.figure.cursor(useblit=True, color='red', linewidth=1)
        pts = []
        pts = np.asarray(self.canvas.figure.ginput(4, timeout=-1))

        xm1 = pts[0][0]
        ym1 = pts[0][1]
        xm2 = pts[1][0]
        ym2 = pts[1][1]

        xm3 = pts[2][0]
        ym3 = pts[2][1]
        xm4 = pts[3][0]
        ym4 = pts[3][1]

        distance1 = ((xm2 - xm1) * (xm2 - xm1) + (ym2 - ym1) * (ym2 - ym1))**0.5
        distance2 = ((xm4 - xm3) * (xm4 - xm3) + (ym4 - ym3) * (ym4 - ym3)) ** 0.5

        if distance1 >= distance2:
            self.Distance = distance1 / self.CoefPxMM
            self.canvas.figure.gca().add_artist(patches.mlines.Line2D([xm1, xm2], [ym1, ym2], color='yellow', linestyle='dotted', linewidth=0.3))
            dist_mm = distance_mm(str(self.MammoCourante), xm1, ym1, xm2, ym2)
        else:
            self.Distance = distance2 / self.CoefPxMM
            self.canvas.figure.gca().add_artist(patches.mlines.Line2D([xm3, xm4], [ym3, ym4], color='yellow', linestyle='dotted', linewidth=0.3))
            dist_mm = distance_mm(str(self.MammoCourante), xm3, ym3, xm4, ym4)

        self.canvas.draw()
        self.UpdateExcel()
        #dist_mm=distance_mm(str(self.MammoCourante),100,200,100,300)
        #self.MesureMGlabel.setText('Mesure MG: ' + str(self.Distance))
        self.MesureMGlabel.setText('Mesure MG: ' + str(dist_mm) + ' mm')

    def SaveButton(self):
        x1 = self.canvas.figure.axes[0].viewLim._points[0, 0]
        y1 = self.canvas.figure.axes[0].viewLim._points[1, 1]
        y2 = self.canvas.figure.axes[0].viewLim._points[0, 1]
        x2 = self.canvas.figure.axes[0].viewLim._points[1, 0]
        #self.label_mamelon_x1.setText('x1: ' + str(x1))
        #self.label_mamelon_y1.setText('y1: ' + str(y1))
        #self.label_mamelon_x2.setText('x2: ' + str(x2))
        #self.label_mamelon_y2.setText('y2: ' + str(y2))
        if self.CoordMamelon == []:
            self.canvas.figure.gca().add_artist(patches.Rectangle((x1, y1), x2 - x1, y2 - y1, edgecolor='blue', fill=False, linestyle='dashdot', linewidth=0.3,zorder=1))
            self.canvas.figure.gca().text(x1, y1, 'M', horizontalalignment='left', verticalalignment='top',color='blue', size='3')
            self.CoordMamelon=[x1,x2,y1,y2]

        self.UpdateExcel()



    def UpdateExcel(self):
        xfile = openpyxl.load_workbook(SOURCE_EXCEL, read_only = False)
        sheet = xfile['seno']

        if len(self.CoordAnomalie1) > 0:
            # rect1X1
            sheet['V' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie1[0])
            # rect1Y1
            sheet['W' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie1[1])
            # rect1X2
            sheet['X' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie1[2])
            # rect1Y2
            sheet['Y' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie1[3])

        if len(self.CoordAnomalie1) == 0:
            # rect1X1
            sheet['V' + str(self.CurseurSourceExcel + 1)] = 0
            # rect1Y1
            sheet['W' + str(self.CurseurSourceExcel + 1)] = 0
            # rect1X2
            sheet['X' + str(self.CurseurSourceExcel + 1)] = 0
            # rect1Y2
            sheet['Y' + str(self.CurseurSourceExcel + 1)] = 0

        if len(self.CoordAnomalie2) > 0:
            # rect2Y1
            sheet['Z' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie2[0])
            # rect2X2
            sheet['AA' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie2[1])
            # rect2Y2
            sheet['AB' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie2[2])
            # rect3X1
            sheet['AC' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie2[3])

        if len(self.CoordAnomalie2) == 0:
            # rect2Y1
            sheet['Z' + str(self.CurseurSourceExcel + 1)] = 0
            # rect2X2
            sheet['AA' + str(self.CurseurSourceExcel + 1)] = 0
            # rect2Y2
            sheet['AB' + str(self.CurseurSourceExcel + 1)] = 0
            # rect3X1
            sheet['AC' + str(self.CurseurSourceExcel + 1)] = 0

        if len(self.CoordAnomalie3) > 0:
            # rect3X1
            sheet['AD' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie3[0])
            # rect3Y1
            sheet['AE' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie3[1])
            # rect3X2
            sheet['Af' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie3[2])
            # rect3Y2
            sheet['AG' + str(self.CurseurSourceExcel + 1)] = int(self.CoordAnomalie3[3])

        if len(self.CoordAnomalie3) == 0:
            # rect3X1
            sheet['AD' + str(self.CurseurSourceExcel + 1)] = 0
            # rect3Y1
            sheet['AE' + str(self.CurseurSourceExcel + 1)] = 0
            # rect3X2
            sheet['Af' + str(self.CurseurSourceExcel + 1)] = 0
            # rect3Y2
            sheet['AG' + str(self.CurseurSourceExcel + 1)] = 0

        if len(self.CoordMamelon) > 0:
            # mamX1
            sheet['AH' + str(self.CurseurSourceExcel + 1)] = int(self.CoordMamelon[0])
            # mamY1
            sheet['AI' + str(self.CurseurSourceExcel + 1)] = int(self.CoordMamelon[1])
            # mamX2
            sheet['AJ' + str(self.CurseurSourceExcel + 1)] = int(self.CoordMamelon[2])
            # mamY2
            sheet['AK' + str(self.CurseurSourceExcel + 1)] = int(self.CoordMamelon[3])

        if len(self.CoordMamelon) == 0:
            # mamX1
            sheet['AH' + str(self.CurseurSourceExcel + 1)] = 0
            # mamY1
            sheet['AI' + str(self.CurseurSourceExcel + 1)] = 0
            # mamX2
            sheet['AJ' + str(self.CurseurSourceExcel + 1)] = 0
            # mamY2
            sheet['AK' + str(self.CurseurSourceExcel + 1)] = 0

        sheet['AL' + str(self.CurseurSourceExcel + 1)] = self.Distance

        self.MesureMGlabel.setText('Mesure MG: ' + str(int(self.Distance)) + ' px')

        xfile.save(SOURCE_EXCEL)

        # # Initialisation après sauvegarde
        # self.CoordMamelon = []
        # self.CoordAnomalie1 = []
        # self.CoordAnomalie2 = []
        # self.CoordAnomalie3 = []

    def UpdateValuesExcel(self):
        xfile = openpyxl.load_workbook(SOURCE_EXCEL, read_only=False)
        sheet = xfile['seno']

        # Côté
        if self.radioDroitButton.isChecked():
            sheet['L' + str(self.CurseurSourceExcel + 1)] = 'R'
        if self.radioGaucheButton.isChecked():
            sheet['L' + str(self.CurseurSourceExcel + 1)] = 'L'

        # Anomalie(s) visible(s)
        if self.AnomaliesVisiblescheckBox.isChecked():
            sheet['E' + str(self.CurseurSourceExcel + 1)] = '1.0'
        else:
            sheet['E' + str(self.CurseurSourceExcel + 1)] = '0'

        # Densité
        if self.radioAButton.isChecked():
            sheet['U' + str(self.CurseurSourceExcel + 1)] = 'A'
        if self.radioBButton.isChecked():
            sheet['U' + str(self.CurseurSourceExcel + 1)] = 'B'
        if self.radioCButton.isChecked():
            sheet['U' + str(self.CurseurSourceExcel + 1)] = 'C'
        if self.radioDButton.isChecked():
            sheet['U' + str(self.CurseurSourceExcel + 1)] = 'D'

        # Anomalie(s)
        if self.MassecheckBox.isChecked():
            sheet['G' + str(self.CurseurSourceExcel + 1)] = '1.0'
        else:
            sheet['G' + str(self.CurseurSourceExcel + 1)] = '0'

        if self.MicroCalcificationscheckBox.isChecked():
            sheet['H' + str(self.CurseurSourceExcel + 1)] = '1.0'
        else:
            sheet['H' + str(self.CurseurSourceExcel + 1)] = '0'

        if self.DesorganisationcheckBox.isChecked():
            sheet['I' + str(self.CurseurSourceExcel + 1)] = '1.0'
        else:
            sheet['I' + str(self.CurseurSourceExcel + 1)] = '0'

        if self.AsymetriecheckBox.isChecked():
            sheet['J' + str(self.CurseurSourceExcel + 1)] = '1.0'
        else:
            sheet['J' + str(self.CurseurSourceExcel + 1)] = '0'

        # Localisation
        if self.SuperieurcheckBox.isChecked():
            sheet['N' + str(self.CurseurSourceExcel + 1)] = '1.0'
        else:
            sheet['N' + str(self.CurseurSourceExcel + 1)] = '0'

        if self.InferieurcheckBox.isChecked():
            sheet['O' + str(self.CurseurSourceExcel + 1)] = '1.0'
        else:
            sheet['O' + str(self.CurseurSourceExcel + 1)] = '0'

        if self.ExternecheckBox.isChecked():
            sheet['P' + str(self.CurseurSourceExcel + 1)] = '1.0'
        else:
            sheet['P' + str(self.CurseurSourceExcel + 1)] = '0'

        if self.InternecheckBox.isChecked():
            sheet['Q' + str(self.CurseurSourceExcel + 1)] = '1.0'
        else:
            sheet['Q' + str(self.CurseurSourceExcel + 1)] = '0'

        if self.retromamcheckBox.isChecked():
            sheet['R' + str(self.CurseurSourceExcel + 1)] = '1.0'
        else:
            sheet['R' + str(self.CurseurSourceExcel + 1)] = '0'




        xfile.save(SOURCE_EXCEL)



if __name__ == '__main__':

    QtWidgets.QApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling, True)
    app = QtWidgets.QApplication(sys.argv)

    Mammo = Figure(dpi=300)

    #
    # fig = plt.figure(figsize=(10, 30), dpi=300)  ## dimension de l'image et résolution en dpi
    # fig.patch.set_facecolor('xkcd:black')  ## choix de la couleur de fond
    # plt.axis('off')  ## suppression des axes et annotations
    # plt.subplots_adjust(left=0, right=1, top=1, bottom=0, wspace=0, hspace=0)  ## réduction des marges sup, inf, dr, g
    #

    left = 0.1
    bottom = 0.05
    width = 0.9
    height = 0.9

    #axe = Mammo.add_axes([left, bottom, width, height])
    axe = Mammo.add_axes([0, 0, 1, 1])

    filename = "IM1.dcm"
    dataset = pydicom.dcmread(filename)
    level = dataset[0x0028, 0x1050][1]
    window = dataset[0x0028, 0x1051][1]
    vmin = level - window
    vmax = level + window

    axe.imshow(dataset.pixel_array, cmap="gray", vmin=vmin, vmax=vmax)

    axe.figure.subplots_adjust(left=0, right=1, top=1, bottom=0, wspace=0, hspace=0)
    axe.figure.patch.set_facecolor('xkcd:black')


    main = Main()
    main.loadImage(Mammo)


    main.show()
    sys.exit(app.exec_())

